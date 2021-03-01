import openpyxl

wb = openpyxl.load_workbook('売上報告書.xlsx', data_only=True)
sheet = wb['売上データ']

sales_list_by_branch = []
sales_by_a, sales_by_b, sales_by_c = 0, 0, 0

for i, cell in enumerate(list(sheet.columns)[2]):
    if cell.value == "A支店":
        sales_by_a += sheet.cell(row=i + 1, column=8).value
    elif cell.value == "B支店":
        sales_by_b += sheet.cell(row=i + 1, column=8).value
    elif cell.value == "C支店":
        sales_by_c += sheet.cell(row=i + 1, column=8).value

# 新しいExcelシートに支店別の売上実績を記載する
sheet_2 = wb.create_sheet('支店別売上実績')
sheet_2['A1'].value, sheet_2['B1'].value = 'A支店', sales_by_a
sheet_2['A2'].value, sheet_2['B2'].value = 'B支店', sales_by_b
sheet_2['A3'].value, sheet_2['B3'].value = 'C支店', sales_by_c
sheet_2['A4'].value, sheet_2['B4'].value = '全支店', sales_by_a + sales_by_b + sales_by_c

wb.save('支店別売上報告書.xlsx')
